# openpyxl should come installed with Anaconda
from openpyxl import load_workbook
import csv


def similarCheck():
    wb = load_workbook(filename='test input file.xlsx') # Change file name here
    ws = wb['Sheet1'] # Assuming the matrix is on the first sheet

    with open('Test output.csv', 'w', newline='') as outcsv:
        writer = csv.writer(outcsv)
        # Creates header in new file
        writer.writerow(['Match %', 'Gene (A)', 'Assay ID (A)',
                         'Gene (B)', 'Assay ID (B)'])
        
        """ Iterates through each cell in .xlsx checking for values in desired range,
            and writes combinations and match % to separate .csv. Duplicate 
            combinations are not checked, cutting processing time down."""            
        for row in ws.rows:
            col = 0
            for cell in row:
                col += 1
                if cell.value == 1:
                    break
                elif type(cell.value) == float:
                    if 0.9 <= cell.value < 1:
                        writer.writerow([cell.value, row[0].value, row[1].value,
                                         ws.cell(row=1, column=col).value, ws.cell(row=2, column=col).value])


similarCheck()
